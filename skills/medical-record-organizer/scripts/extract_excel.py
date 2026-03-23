#!/usr/bin/env python3
"""
extract_excel.py — Extract data from .xlsx Excel files.

Usage:
    python3 extract_excel.py PATH

Output JSON:
    {"success": true, "text": "...", "sheets": ["Sheet1"], "total_rows": 25}
    {"success": false, "error": "..."}
"""

import json
import sys
from pathlib import Path


def extract_xlsx(file_path: Path) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "error": "openpyxl not installed. Run: uv pip install openpyxl"}

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheet_names = wb.sheetnames
    all_parts = []
    total_rows = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        sheet_rows = []
        for row in ws.iter_rows(values_only=True):
            # Skip entirely empty rows
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                sheet_rows.append("\t".join(cells))
                total_rows += 1
        if sheet_rows:
            all_parts.append(f"[Sheet: {sheet_name}]")
            all_parts.extend(sheet_rows)

    full_text = "\n".join(all_parts)
    return {
        "success": True,
        "text": full_text,
        "sheets": sheet_names,
        "total_rows": total_rows,
    }


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "Usage: python3 extract_excel.py PATH"}))
        sys.exit(1)

    file_path = Path(sys.argv[1]).expanduser().resolve()

    if not file_path.exists():
        print(json.dumps({"success": False, "error": f"File not found: {file_path}"}))
        sys.exit(1)

    suffix = file_path.suffix.lower()
    if suffix == ".xls":
        print(json.dumps({
            "success": False,
            "error": "旧版 .xls 格式不支持直接提取。请在 Excel 中另存为 .xlsx 后重试。",
        }))
        sys.exit(1)

    if suffix != ".xlsx":
        print(json.dumps({"success": False, "error": f"不支持的文件格式: {suffix}，仅支持 .xlsx"}))
        sys.exit(1)

    result = extract_xlsx(file_path)
    print(json.dumps(result, ensure_ascii=False))
    if not result["success"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
